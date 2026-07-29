from __future__ import annotations

import io
import json
from typing import Any

from openpyxl import load_workbook


def apply_cell_updates(
    content: bytes,
    updates: list[dict[str, Any]],
) -> bytes:
    """
    Apply per-row cell updates to an xlsx workbook.

    Each update:
      {
        "sheet": "C",          # sheet name
        "row": 2,              # 1-based Excel row (header is usually 1)
        "values": {"id": "...", "managingOrganization": "..."}
      }

    Missing target columns (e.g. `id` on FHIR templates) are created as new header columns.
    """
    book = load_workbook(io.BytesIO(content))
    header_maps: dict[str, dict[str, int]] = {}

    for update in updates:
        sheet_name = str(update["sheet"])
        if sheet_name not in book.sheetnames:
            continue
        sheet = book[sheet_name]
        if sheet_name not in header_maps:
            header_maps[sheet_name] = _header_map(sheet)
        header_map = header_maps[sheet_name]
        excel_row = int(update["row"])
        values: dict[str, Any] = update.get("values") or {}
        for column_name, value in values.items():
            col_idx = header_map.get(column_name)
            if col_idx is None:
                for header, idx in header_map.items():
                    if header.casefold() == column_name.casefold():
                        col_idx = idx
                        break
            if col_idx is None:
                col_idx = _ensure_column(sheet, header_map, column_name)
            sheet.cell(row=excel_row, column=col_idx, value=_cell_value(value))

    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


def _header_map(sheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in sheet[1]:
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        if name:
            mapping[name] = cell.column
    return mapping


def _ensure_column(sheet, header_map: dict[str, int], column_name: str) -> int:
    """Append a new header column and register it in header_map."""
    next_col = (max(header_map.values()) + 1) if header_map else 1
    sheet.cell(row=1, column=next_col, value=column_name)
    header_map[column_name] = next_col
    return next_col


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def organization_reference(org_id: str) -> dict[str, str]:
    return {"reference": f"Organization/{org_id}"}


def location_reference(location_id: str) -> dict[str, str]:
    return {"reference": f"Location/{location_id}"}
